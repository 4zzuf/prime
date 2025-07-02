"""Calcula presupuestos y necesidades para sistemas fotovoltaicos."""

from __future__ import annotations

import os
from typing import Dict, List, Tuple
import math
import re

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover - dependency may not be installed
    Workbook = None  # type: ignore
    load_workbook = None  # type: ignore

FILE = "equipos.xlsx"
LOADS_FILE = "cargas.xlsx"
SHEETS = ["Paneles", "Inversores", "Baterias", "Controladores"]
CATEGORIES = ["Barato", "Intermedio", "Premium"]


def crear_excel_de_ejemplo(filename: str) -> None:
    """Crea un archivo Excel con datos de ejemplo si no existe."""

    if Workbook is None:
        raise ImportError("openpyxl no esta instalado")

    wb = Workbook()
    # Elimina la hoja predeterminada
    wb.remove(wb.active)

    datos = {
        "Paneles": [
            ("Barato", "Eco20", "20W", 60),
            ("Barato", "Eco30", "30W", 70),
            ("Barato", "Eco50", "50W", 100),
            ("Intermedio", "Mid100", "100W", 180),
            ("Intermedio", "Mid150", "150W", 230),
            ("Intermedio", "Mid200", "200W", 280),
            ("Premium", "Pro605", "605W", 800),
            # Solo se consideran opciones de 100W en adelante
            ("Barato", "Basic100", "100W", 150),
            ("Intermedio", "Mid200", "200W", 250),
            ("Premium", "Top300", "300W", 400),
        ],
        "Inversores": [
            ("Barato", "Mod500", "Onda modificada 500W", 150),
            ("Barato", "Mod1000", "Onda modificada 1000W", 200),
            ("Intermedio", "Pura1500", "Onda pura 1500W", 350),
            ("Intermedio", "Pura1800", "Onda pura 1800W", 420),
            ("Premium", "Hibrido4000", "Híbrido 4000W", 900),
            ("Premium", "Hibrido3000", "Híbrido 3000W", 700),
        ],
        "Baterias": [
            ("Barato", "AGM7", "AGM 7Ah", 12, 40),
            ("Barato", "AGM26", "AGM 26Ah", 12, 70),
            ("Barato", "AGM40", "AGM 40Ah", 12, 100),
            ("Barato", "AGM75", "AGM 75Ah", 12, 150),
            ("Intermedio", "Gel100", "Gel 100Ah", 12, 250),
            ("Intermedio", "Gel150", "Gel 150Ah", 12, 320),
            ("Intermedio", "Gel200", "Gel 200Ah", 24, 400),
            ("Intermedio", "Gel300", "Gel 300Ah", 48, 600),
            ("Premium", "Li100", "Litio 100Ah", 12, 800),
            ("Premium", "Li200", "Litio 200Ah", 24, 1400),
        ],
        "Controladores": [
            ("Barato", "PWM10", "PWM 10A", 30),
            ("Barato", "PWM20", "PWM 20A", 45),
            ("Barato", "PWM30", "PWM 30A", 60),
            ("Intermedio", "MPPT15", "MPPT 15A", 120),
            ("Premium", "MPPT20", "MPPT 20A", 180),
            ("Barato", "PWM10", "PWM 10A", 50),
            ("Intermedio", "PWM20", "PWM 20A", 100),
            ("Premium", "MPPT30", "MPPT 30A", 150),

        ],
    }

    for nombre, filas in datos.items():
        ws = wb.create_sheet(title=nombre)
        if nombre == "Baterias":
            ws.append(["Categoria", "Marca", "Detalle", "Voltaje", "Precio"])
        else:
            ws.append(["Categoria", "Marca", "Detalle", "Precio"])

        for fila in filas:
            ws.append(fila)

    wb.save(filename)


def crear_excel_cargas_de_ejemplo(filename: str) -> None:
    """Crea un archivo Excel con cargas de ejemplo."""

    if Workbook is None:
        raise ImportError("openpyxl no esta instalado")

    wb = Workbook()
    ws = wb.active
    ws.title = "Cargas"

    # Encabezado en dos filas para coincidir con el formato solicitado
    ws.append([
        "Nombre del Aparato",
        "Cantidad",
        "Carga (W)",
        "Uso",
        "",
    ])
    ws.append(["", "", "", "Horas por día", "Horas noche"])

    datos = [
        ("Celular", 0, "10 W", "0 hr", "2 hr"),
        ("Lavadora", 0, "300 W", "2 hr", "1 hr"),
        ("TV pequeña 32''", 2, "60 W", "2 hr", "3 hr"),
        ("Licuadora", 1, "400 W", "0.083333333 hr", "0 hr"),
        ("Calenton electrico", 0, "1000 W", "4 hr", "7 hr"),
        ("Nevera", 0, "400 W", "6 hr", "6 hr"),
        ("Refrigerador grande", 0, "250 W", "5 hr", "5 hr"),
        ("Focos", 5, "5 W", "0 hr", "4 hr"),
        ("Horno", 0, "200 W", "5 hr", "7 hr"),
        ("Computadora", 0, "200 W", "3 hr", "3 hr"),
        ("Ducha electrica", 0, "200 W", "5 hr", "7 hr"),
        ("Laptop ", 0, "60 W", "2 hr", "3 hr"),
        ("Radio", 0, "20 W", "2 hr", "5 hr"),
        ("starlink", 0, "63 W", "10 hr", "10 hr"),
        ("Camaras", 0, "15 W", "24 hr", "7 hr"),
        ("Tv grande", 0, "150 W", "3 hr", "3 hr"),
        ("Refri pequeña", 1, "100 W", "6 hr", "6 hr"),
        ("Parlante bletooth", 0, "20 W", "1 hr", "7 hr"),
        ("Radio pequeña", 0, "5 W", "4 hr", "3 hr"),
    ]
    for fila in datos:
        ws.append(fila)

    wb.save(filename)

def _extraer_numero(texto: str) -> float:
    """Extrae el primer numero encontrado en un texto."""

    m = re.search(r"([0-9]+(?:\.[0-9]+)?)", texto)
    return float(m.group(1)) if m else 0.0


def leer_datos(filename: str) -> Dict[str, Dict[str, List[Tuple[str, float, float]]]]:
    """Lee el archivo Excel y organiza los datos por componente y categoria.

    Devuelve nombre, capacidad (W, Ah o A) y precio.
    """

    if load_workbook is None:
        raise ImportError("openpyxl no esta instalado")

    wb = load_workbook(filename)
    datos: Dict[str, Dict[str, List[Tuple]]] = {}
    for hoja in SHEETS:
        ws = wb[hoja]
        datos[hoja] = {cat: [] for cat in CATEGORIES}

        for row in ws.iter_rows(min_row=2, values_only=True):
            if hoja == "Baterias":
                if len(row) < 5:
                    continue
                categoria, marca, detalle, volt, precio = row[:5]
                if categoria not in CATEGORIES:
                    continue
                nombre = f"{marca} {detalle}"
                capacidad = _extraer_numero(str(detalle))
                datos[hoja][categoria].append(
                    (nombre, capacidad, float(volt), float(precio))
                )
            else:
                if len(row) < 4:
                    continue
                categoria, marca, detalle, precio = row
                if categoria in CATEGORIES:
                    nombre = f"{marca} {detalle}"
                    capacidad = _extraer_numero(str(detalle))
                    datos[hoja][categoria].append((nombre, capacidad, float(precio)))
    return datos

def leer_cargas(filename: str) -> List[Dict[str, float]]:
    """Lee el excel de cargas y devuelve una lista de diccionarios."""

    if load_workbook is None:
        raise ImportError("openpyxl no esta instalado")

    wb = load_workbook(filename)
    ws = wb.active
    cargas = []
    # El archivo de ejemplo utiliza dos filas de encabezado
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row:
            continue

        valores = list(row)
        if len(valores) < 5:
            valores.extend([0.0] * (5 - len(valores)))

        (
            aparato,
            cantidad,
            carga,
            horas_dia,
            horas_noche,
        ) = valores[:5]

        cargas.append(
            {
                "aparato": str(aparato),
                "cantidad": _extraer_numero(str(cantidad)),
                "carga": _extraer_numero(str(carga)),
                "horas_dia": _extraer_numero(str(horas_dia)),
                "horas_noche": _extraer_numero(str(horas_noche)),

            }
        )
    return cargas


def seleccionar_cargas_gui(cargas: List[Dict[str, float]]) -> List[Dict[str, float]]:
    """Muestra una ventana (PyQt5) para elegir y editar las cargas."""

    try:
        from PyQt5 import QtCore, QtWidgets
    except Exception as exc:  # pragma: no cover - dependencias ausentes
        print(f"No se pudo abrir la interfaz grafica: {exc}")
        return cargas

    app = QtWidgets.QApplication([])
    dialog = QtWidgets.QDialog()
    dialog.setWindowTitle("Seleccionar cargas")
    layout = QtWidgets.QVBoxLayout(dialog)

    headers = [
        "Usar",
        "Aparato",
        "Cantidad",
        "Carga(W)",
        "HorasDia",
        "HorasNoche",
    ]
    table = QtWidgets.QTableWidget(len(cargas), len(headers))
    table.setHorizontalHeaderLabels(headers)
    dialog.resize(1100, 700)

    for row, carga in enumerate(cargas):
        chk_item = QtWidgets.QTableWidgetItem()
        chk_item.setCheckState(QtCore.Qt.Checked)
        table.setItem(row, 0, chk_item)
        table.setItem(row, 1, QtWidgets.QTableWidgetItem(carga["aparato"]))
        table.setItem(row, 2, QtWidgets.QTableWidgetItem(str(carga["cantidad"])))
        table.setItem(row, 3, QtWidgets.QTableWidgetItem(str(carga["carga"])))
        table.setItem(row, 4, QtWidgets.QTableWidgetItem(str(carga["horas_dia"])))
        table.setItem(row, 5, QtWidgets.QTableWidgetItem(str(carga["horas_noche"])))

    layout.addWidget(table)
    boton = QtWidgets.QPushButton("Calcular")
    layout.addWidget(boton)

    resultado: List[Dict[str, float]] = []

    def finalizar() -> None:
        for row in range(table.rowCount()):
            item_usar = table.item(row, 0)
            if item_usar.checkState() != QtCore.Qt.Checked:
                continue
            aparato = table.item(row, 1).text()
            cantidad = float(table.item(row, 2).text() or 0)
            carga_w = float(table.item(row, 3).text() or 0)
            horas_dia = float(table.item(row, 4).text() or 0)
            horas_noche = float(table.item(row, 5).text() or 0)
            resultado.append(
                {
                    "aparato": aparato,
                    "cantidad": cantidad,
                    "carga": carga_w,
                    "horas_dia": horas_dia,
                    "horas_noche": horas_noche,
                }
            )
        dialog.accept()

    boton.clicked.connect(finalizar)
    dialog.exec_()
    app.quit()
    return resultado or cargas


def curva_irradiacion_cusco() -> Dict[int, float]:
    """Devuelve una curva horaria de irradiación típica de Cusco."""

    # Valores aproximados en W/m^2 de 6 AM a 6 PM
    valores = [
        0,
        100,
        300,
        500,
        700,
        850,
        950,
        1000,
        950,
        800,
        600,
        400,
        200,
        0,
    ]
    horas = list(range(6, 20))
    return dict(zip(horas, valores))


def horas_solares_efectivas(curva: Dict[int, float]) -> float:
    """Devuelve horas solares pico aproximadas.

    Se simplifica a un valor fijo de 5 horas para evitar sobreestimar la
    generación solar.
    """

    # Aunque podría calcularse a partir de la curva, se fija en ~5 h para
    # reflejar condiciones más conservadoras.
    return 5.0


def energia_dia_noche(
    cargas: List[Dict[str, float]], curva: Dict[int, float]
) -> Tuple[float, float]:
    """Calcula energia consumida de dia y de noche."""

    energia_dia = 0.0
    energia_noche = 0.0

    for carga in cargas:
        potencia = carga["carga"] * carga["cantidad"]
        energia_dia += potencia * carga.get("horas_dia", 0)
        energia_noche += potencia * carga.get("horas_noche", 0)

    return energia_dia, energia_noche


def calcular_necesidades(
    cargas: List[Dict[str, float]], curva: Dict[int, float]
) -> Tuple[float, float]:
    """Calcula potencia de panel y capacidad de bateria necesarias."""

    energia_dia, energia_noche = energia_dia_noche(cargas, curva)
    hs = horas_solares_efectivas(curva)
    potencia_panel = (energia_dia + energia_noche) / hs if hs else 0
    capacidad_bateria = energia_noche / 12  # Ah para bateria de 12V
    return potencia_panel, capacidad_bateria


def potencia_maxima_demanda(cargas: List[Dict[str, float]]) -> float:
    """Calcula la potencia simultanea maxima de las cargas."""

    return sum(carga["carga"] * carga["cantidad"] for carga in cargas)


def calcular_kit(
    datos: Dict[str, Dict[str, List[Tuple]]],
    potencia_panel: float,
    capacidad_bateria: float,
    demanda_maxima: float,
) -> Dict[str, Dict[str, Tuple[str, float]]]:
    """Selecciona componentes suficientes para cubrir las necesidades."""

    resultados: Dict[str, Dict[str, Tuple[str, float]]] = {
        cat: {} for cat in CATEGORIES
    }

    if potencia_panel <= 1500:
        volt_sistema = 12
    elif potencia_panel <= 5000:
        volt_sistema = 24
    else:
        volt_sistema = 48

    energia_noche_kwh = capacidad_bateria * 12 / 1000

    for categoria in CATEGORIES:
        # Paneles
        mejor_total = math.inf
        mejor_desc = "Sin datos"
        for nombre, capacidad, precio in datos["Paneles"].get(categoria, []):
            if capacidad <= 0:
                continue
            cantidad = math.ceil(potencia_panel / capacidad)
            total = cantidad * precio
            if total < mejor_total:
                mejor_total = total
                mejor_desc = f"{cantidad} x {nombre}"
        resultados[categoria]["Paneles"] = (mejor_desc, mejor_total if mejor_total < math.inf else 0.0)

        # Baterias considerando DoD y voltaje de sistema
        mejor_total = math.inf
        mejor_desc = "Sin datos"
        dod = 0.5 if categoria in ("Barato", "Intermedio") else 0.9
        ah_sistema = energia_noche_kwh * 1000 / volt_sistema
        capacidad_requerida = ah_sistema / dod if dod else ah_sistema
        for nombre, capacidad, volt, precio in datos["Baterias"].get(categoria, []):
            if capacidad <= 0 or volt_sistema % volt != 0:
                continue
            en_serie = int(volt_sistema / volt)
            en_paralelo = math.ceil(capacidad_requerida / capacidad)
            total_bat = en_serie * en_paralelo
            total = total_bat * precio
            desc = f"{total_bat} x {nombre} ({en_serie}S{en_paralelo}P)"
            if total < mejor_total:
                mejor_total = total
                mejor_desc = desc
        resultados[categoria]["Baterias"] = (
            mejor_desc,
            mejor_total if mejor_total < math.inf else 0.0,
        )

        # Inversores
        mejor_precio = math.inf
        mejor_desc = "Sin datos"
        for nombre, capacidad, precio in datos["Inversores"].get(categoria, []):
            if capacidad >= demanda_maxima and precio < mejor_precio:
                mejor_precio = precio
                mejor_desc = nombre
        resultados[categoria]["Inversores"] = (mejor_desc, mejor_precio if mejor_precio < math.inf else 0.0)

        # Controladores: se elige el mas barato
        mejor_precio = math.inf
        mejor_desc = "Sin datos"
        for nombre, capacidad, precio in datos["Controladores"].get(categoria, []):
            if precio < mejor_precio:
                mejor_precio = precio
                mejor_desc = nombre
        resultados[categoria]["Controladores"] = (mejor_desc, mejor_precio if mejor_precio < math.inf else 0.0)

    return resultados


def imprimir_presupuestos(presupuestos: Dict[str, Dict[str, Tuple[str, float]]]) -> None:
    """Muestra los presupuestos y los componentes utilizados."""

    for categoria in CATEGORIES:
        elementos = presupuestos[categoria]
        total = sum(precio for _, precio in elementos.values())
        print(f"{categoria}: ${total:.2f}")
        for componente, (marca, precio) in elementos.items():
            print(f"  {componente}: {marca} - ${precio:.2f}")
        print()


def main() -> None:
    if not os.path.exists(FILE):
        crear_excel_de_ejemplo(FILE)
        print(f"Se creó el archivo '{FILE}' con datos de ejemplo.")

    if not os.path.exists(LOADS_FILE):
        crear_excel_cargas_de_ejemplo(LOADS_FILE)
        print(f"Se creó el archivo '{LOADS_FILE}' con datos de ejemplo.")

    datos = leer_datos(FILE)
    cargas = leer_cargas(LOADS_FILE)
    curva = curva_irradiacion_cusco()
    potencia_panel, capacidad_bateria = calcular_necesidades(cargas, curva)
    demanda_maxima = potencia_maxima_demanda(cargas)
    presupuestos = calcular_kit(datos, potencia_panel, capacidad_bateria, demanda_maxima)

    imprimir_presupuestos(presupuestos)
    print("Requerimientos del sistema:")
    print(f"  Potencia de panel requerida: {potencia_panel:.2f} W")
    print(f"  Capacidad de batería requerida: {capacidad_bateria:.2f} Ah")


if __name__ == "__main__":
    main()
