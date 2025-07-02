"""Compara costo de red electrica vs sistema solar y genera graficos."""

from __future__ import annotations

import os
import re
import hashlib
from typing import Dict, Tuple

from Precios import (
    FILE,
    LOADS_FILE,
    CATEGORIES,
    INVENTARIO_FILE,
    INGRESOS_FILE,
    crear_excel_de_ejemplo,
    crear_excel_cargas_de_ejemplo,
    crear_excel_inventario,
    crear_excel_ingresos,
    leer_datos,
    leer_cargas,
    leer_inventario,
    guardar_inventario,
    registrar_movimiento,
    curva_irradiacion_cusco,
    calcular_necesidades,
    energia_dia_noche,
    potencia_maxima_demanda,
    calcular_kit,
    seleccionar_cargas_gui,
)

COSTO_RED = 0.83  # PEN por kWh
VIDA_UTIL_ANIOS = 20

# --- Credenciales para modificar inventario ---
LOGIN_USER = "Michifus"
_SALT = b"\xf9\x1d%T\xd9\x96\xc5\xf7\xca?2\xaa\x81\xb1`L"
_PWD_HASH = "9628bfacb991b822576cf52d147609f0146ef646b3efbb71372517d1af07db14"

def _verificar_login(usuario: str, contrasena: str) -> bool:
    """Devuelve True si las credenciales coinciden."""
    hashed = hashlib.pbkdf2_hmac(
        "sha256", contrasena.encode(), _SALT, 100000
    ).hex()
    return usuario == LOGIN_USER and hashed == _PWD_HASH

def energia_diaria_kwh(cargas: list[dict[str, float]], curva: Dict[int, float]) -> float:
    """Suma el consumo diario en kWh a partir de los intervalos."""

    energia_dia, energia_noche = energia_dia_noche(cargas, curva)
    return (energia_dia + energia_noche) / 1000



def calcular_amortizacion(
    presupuesto: Dict[str, Tuple[str, float]], daily_kwh: float
) -> Tuple[float, float, float, float]:
    """Devuelve costo del sistema, costo por kWh, payback y ahorro."""

    costo_sistema = sum(p for _, p in presupuesto.values())
    costo_anual_red = daily_kwh * COSTO_RED * 365
    payback = costo_sistema / costo_anual_red if costo_anual_red else float("inf")
    costo_kwh = (
        costo_sistema / (daily_kwh * 365 * VIDA_UTIL_ANIOS)
        if daily_kwh
        else float("inf")
    )
    ahorro_total = costo_anual_red * VIDA_UTIL_ANIOS - costo_sistema
    return costo_sistema, costo_kwh, payback, ahorro_total


def graficar_costo_acumulado(costo_sistema: float, daily_kwh: float, nombre: str) -> None:
    """Genera un grafico de costo acumulado y lo guarda."""

    try:
        import matplotlib.pyplot as plt
    except Exception as exc:  # pragma: no cover - si matplotlib no esta
        raise ImportError("matplotlib no esta instalado") from exc

    anios = list(range(VIDA_UTIL_ANIOS + 1))
    costo_red = [daily_kwh * COSTO_RED * 365 * a for a in anios]
    costo_solar = [costo_sistema if a > 0 else 0 for a in anios]

    plt.figure()
    plt.plot(anios, costo_red, label="Red electrica")
    plt.plot(anios, costo_solar, label="Sistema solar")
    plt.xlabel("Años")
    plt.ylabel("PEN")
    plt.title(f"Costo acumulado - {nombre}")
    plt.legend()
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(f"costo_{nombre}.png")
    plt.close()

def graficar_costo_anual(costo_sistema: float, daily_kwh: float, nombre: str) -> None:
    """Grafica el costo anual de red vs el costo anual amortizado del kit."""

    try:
        import matplotlib.pyplot as plt
    except Exception as exc:  # pragma: no cover - si matplotlib no esta
        raise ImportError("matplotlib no esta instalado") from exc

    anios = list(range(1, VIDA_UTIL_ANIOS + 1))
    costo_red = [daily_kwh * COSTO_RED * 365 for _ in anios]
    costo_solar = [costo_sistema / VIDA_UTIL_ANIOS for _ in anios]

    plt.figure()
    plt.plot(anios, costo_red, label="Red electrica")
    plt.plot(anios, costo_solar, label="Sistema solar (amortizado)")
    plt.xlabel("Años")
    plt.ylabel("PEN por año")
    plt.title(f"Costo anual - {nombre}")
    plt.legend()
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(f"costo_anual_{nombre}.png")
    plt.close()


def graficar_ahorro_largo_plazo(costo_sistema: float, daily_kwh: float, nombre: str) -> None:
    """Grafica el ahorro acumulado durante 10 años."""

    try:
        import matplotlib.pyplot as plt
    except Exception as exc:  # pragma: no cover - si matplotlib no esta
        raise ImportError("matplotlib no esta instalado") from exc

    anios = list(range(1, 11))
    costo_red = [daily_kwh * COSTO_RED * 365 * a for a in anios]
    costo_solar = [(costo_sistema / VIDA_UTIL_ANIOS) * a for a in anios]
    ahorro = [cr - cs for cr, cs in zip(costo_red, costo_solar)]

    plt.figure()
    plt.plot(anios, ahorro, label="Ahorro acumulado")
    plt.xlabel("Años")
    plt.ylabel("PEN")
    plt.title(f"Ahorro a largo plazo - {nombre}")
    plt.legend()
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(f"ahorro_{nombre}.png")
    plt.close()
    
def main() -> None:
    """Abre una interfaz grafica para la simulacion."""

    if not os.path.exists(FILE):
        crear_excel_de_ejemplo(FILE)
        print(f"Se creó el archivo '{FILE}' con datos de ejemplo.")

    if not os.path.exists(LOADS_FILE):
        crear_excel_cargas_de_ejemplo(LOADS_FILE)
        print(f"Se creó el archivo '{LOADS_FILE}' con datos de ejemplo.")

    if not os.path.exists(INVENTARIO_FILE):
        crear_excel_inventario(INVENTARIO_FILE)
    if not os.path.exists(INGRESOS_FILE):
        crear_excel_ingresos(INGRESOS_FILE)

    datos = leer_datos(FILE)
    cargas_base = leer_cargas(LOADS_FILE)
    inventario = leer_inventario(INVENTARIO_FILE)

    try:
        from PyQt5 import QtCore, QtGui, QtWidgets
    except Exception as exc:  # pragma: no cover - dependencias ausentes
        print(f"No se pudo abrir la interfaz grafica: {exc}")
        return

    app = QtWidgets.QApplication([])
    ventana = QtWidgets.QWidget()
    ventana.setWindowTitle("Simulador Solar")

    def pedir_login() -> bool:
        """Solicita credenciales y las valida."""
        dlg = QtWidgets.QDialog(ventana)
        dlg.setWindowTitle("Iniciar sesión")
        form = QtWidgets.QFormLayout(dlg)
        usuario = QtWidgets.QLineEdit()
        contrasena = QtWidgets.QLineEdit()
        contrasena.setEchoMode(QtWidgets.QLineEdit.Password)
        form.addRow("Usuario", usuario)
        form.addRow("Contraseña", contrasena)
        botones = QtWidgets.QDialogButtonBox(
            QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel
        )
        form.addRow(botones)

        def aceptar() -> None:
            if _verificar_login(usuario.text(), contrasena.text()):
                dlg.accept()
            else:
                QtWidgets.QMessageBox.warning(
                    dlg, "Error", "Credenciales incorrectas"
                )

        botones.accepted.connect(aceptar)
        botones.rejected.connect(dlg.reject)
        return dlg.exec_() == QtWidgets.QDialog.Accepted
    layout_principal = QtWidgets.QHBoxLayout(ventana)

    headers = ["Usar", "Aparato", "Cantidad", "Carga(W)", "HorasDia", "HorasNoche"]
    tabla = QtWidgets.QTableWidget(len(cargas_base), len(headers))
    tabla.setHorizontalHeaderLabels(headers)
    tabla.setStyleSheet("background-color:#e8f4ff;")
    ventana.resize(1300, 800)
    for fila, carga in enumerate(cargas_base):
        chk = QtWidgets.QTableWidgetItem()
        chk.setCheckState(QtCore.Qt.Checked)
        tabla.setItem(fila, 0, chk)
        tabla.setItem(fila, 1, QtWidgets.QTableWidgetItem(carga["aparato"]))
        tabla.setItem(fila, 2, QtWidgets.QTableWidgetItem(str(carga["cantidad"])))
        tabla.setItem(fila, 3, QtWidgets.QTableWidgetItem(str(carga["carga"])))
        tabla.setItem(fila, 4, QtWidgets.QTableWidgetItem(str(carga["horas_dia"])))
        tabla.setItem(fila, 5, QtWidgets.QTableWidgetItem(str(carga["horas_noche"])))

    # ----- Lado izquierdo -----
    layout_izq = QtWidgets.QVBoxLayout()
    layout_izq.addWidget(tabla)
    btn_toggle = QtWidgets.QPushButton("Marcar/Desmarcar todo")
    layout_izq.addWidget(btn_toggle)
    layout_principal.addLayout(layout_izq)

    # ----- Lado derecho -----
    layout_der = QtWidgets.QVBoxLayout()
    btn_ejecutar = QtWidgets.QPushButton("Ejecutar simulacion")
    layout_der.addWidget(btn_ejecutar)
    btn_costo = QtWidgets.QPushButton("Costo acumulado")
    btn_anual = QtWidgets.QPushButton("Costo anual")
    btn_ahorro = QtWidgets.QPushButton("Ahorro")
    btn_sistemas = QtWidgets.QPushButton("Sistemas")
    btn_inventario = QtWidgets.QPushButton("Inventario")
    btn_movimientos = QtWidgets.QPushButton("Ingresos/Egresos")
    for b in (btn_costo, btn_anual, btn_ahorro, btn_sistemas):
        b.setEnabled(False)
        layout_der.addWidget(b)
    for b in (btn_inventario, btn_movimientos):
        b.setEnabled(True)
        layout_der.addWidget(b)

    salida = QtWidgets.QTextEdit()
    salida.setReadOnly(True)
    salida.setStyleSheet("background-color:#fff9e6;")
    layout_der.addWidget(salida)

    layout_der.addStretch(1)
    layout_principal.addLayout(layout_der)
    def parse_item(desc: str) -> Tuple[str, int]:
        m = re.match(r"(\d+)\s*x\s*([^()]+)", desc)
        if m:
            qty = int(m.group(1))
            name = m.group(2).strip()
        else:
            qty = 1
            name = desc.split("(")[0].strip()
        return name, qty
    def toggle_checks() -> None:
        """Marca o desmarca todas las cargas."""
        any_unchecked = any(
            tabla.item(r, 0).checkState() != QtCore.Qt.Checked
            for r in range(tabla.rowCount())
        )
        nuevo = QtCore.Qt.Checked if any_unchecked else QtCore.Qt.Unchecked
        for r in range(tabla.rowCount()):
            tabla.item(r, 0).setCheckState(nuevo)

    resultados: Dict[str, Dict[str, Tuple[str, float]]] = {}
    daily_kwh: float = 0.0

    def vender_sistema(cat: str, con_igv: bool) -> None:
        pres = resultados.get(cat)
        if not pres:
            return
        total = sum(p for _, p in pres.values())
        concepto = f"Venta {cat} {'con' if con_igv else 'sin'} IGV"
        if con_igv:
            total *= 1.18
        faltantes = []
        for desc, _ in pres.values():
            nombre, cant = parse_item(desc)
            if inventario.get(nombre, 0) < cant:
                faltantes.append(nombre)
        if faltantes:
            QtWidgets.QMessageBox.warning(
                ventana,
                "Stock insuficiente",
                "No hay stock para: " + ", ".join(faltantes),
            )
            return
        for desc, _ in pres.values():
            nombre, cant = parse_item(desc)
            inventario[nombre] = inventario.get(nombre, 0) - cant
        guardar_inventario(INVENTARIO_FILE, inventario)
        registrar_movimiento(INGRESOS_FILE, concepto, total)
        QtWidgets.QMessageBox.information(ventana, "Venta", "Venta registrada")

    def mostrar_inventario() -> None:
        dlg = QtWidgets.QDialog(ventana)
        dlg.setWindowTitle("Inventario")
        lay = QtWidgets.QVBoxLayout(dlg)
        table = QtWidgets.QTableWidget(len(inventario), 2)
        table.setHorizontalHeaderLabels(["Producto", "Cantidad"])
        for r, (prod, cant) in enumerate(inventario.items()):
            table.setItem(r, 0, QtWidgets.QTableWidgetItem(prod))
            table.setItem(r, 1, QtWidgets.QTableWidgetItem(str(cant)))
        table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        lay.addWidget(table)

        buttons = QtWidgets.QHBoxLayout()
        btn_mod = QtWidgets.QPushButton("Modificar inventario")
        btn_cerrar = QtWidgets.QPushButton("Cerrar")
        buttons.addWidget(btn_mod)
        buttons.addWidget(btn_cerrar)
        lay.addLayout(buttons)

        def habilitar_edicion() -> None:
            if not pedir_login():
                return
            table.setEditTriggers(
                QtWidgets.QAbstractItemView.DoubleClicked
                | QtWidgets.QAbstractItemView.EditKeyPressed
            )
            btn_mod.setEnabled(False)
            btn_guardar = QtWidgets.QPushButton("Guardar cambios")
            buttons.addWidget(btn_guardar)

            def guardar() -> None:
                for r in range(table.rowCount()):
                    prod_item = table.item(r, 0)
                    qty_item = table.item(r, 1)
                    if prod_item is None or qty_item is None:
                        continue
                    producto = prod_item.text()
                    try:
                        cantidad = float(qty_item.text())
                        if cantidad < 0:
                            raise ValueError
                    except Exception:
                        QtWidgets.QMessageBox.warning(
                            dlg,
                            "Error",
                            f"Cantidad invalida para '{producto}'",
                        )
                        return
                    inventario[producto] = cantidad
                guardar_inventario(INVENTARIO_FILE, inventario)
                table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
                QtWidgets.QMessageBox.information(
                    dlg, "Inventario", "Cambios guardados"
                )
                btn_guardar.setEnabled(False)

            btn_guardar.clicked.connect(guardar)

        btn_mod.clicked.connect(habilitar_edicion)
        btn_cerrar.clicked.connect(dlg.accept)

        dlg.resize(600, 400)
        dlg.exec_()

    def mostrar_movimientos() -> None:
        dlg = QtWidgets.QDialog(ventana)
        dlg.setWindowTitle("Ingresos y egresos")
        lay = QtWidgets.QVBoxLayout(dlg)
        movs = []
        if os.path.exists(INGRESOS_FILE):
            from openpyxl import load_workbook

            wb = load_workbook(INGRESOS_FILE)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                movs.append(row)
        table = QtWidgets.QTableWidget(len(movs), 3)
        table.setHorizontalHeaderLabels(["Fecha", "Concepto", "Monto"])
        for r, (f, c, m) in enumerate(movs):
            table.setItem(r, 0, QtWidgets.QTableWidgetItem(str(f)))
            table.setItem(r, 1, QtWidgets.QTableWidgetItem(str(c)))
            table.setItem(r, 2, QtWidgets.QTableWidgetItem(f"{m:.2f}"))
        table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        lay.addWidget(table)
        form = QtWidgets.QHBoxLayout()
        txt_c = QtWidgets.QLineEdit()
        txt_m = QtWidgets.QLineEdit()
        btn_add = QtWidgets.QPushButton("Registrar")
        form.addWidget(txt_c)
        form.addWidget(txt_m)
        form.addWidget(btn_add)
        lay.addLayout(form)

        def registrar_local() -> None:
            try:
                monto = float(txt_m.text() or 0)
            except ValueError:
                monto = 0.0
            concepto = txt_c.text() or "Movimiento"
            registrar_movimiento(INGRESOS_FILE, concepto, monto)
            dlg.accept()

        btn_add.clicked.connect(registrar_local)
        dlg.resize(700, 500)
        dlg.exec_()

    def ejecutar() -> None:
        nonlocal resultados, daily_kwh
        cargas: list[dict[str, float]] = []
        # Asegura que se guarden los cambios en celdas editadas
        for r in range(tabla.rowCount()):
            for c in range(tabla.columnCount()):
                item = tabla.item(r, c)
                if item is not None:
                    tabla.closePersistentEditor(item)
        for row in range(tabla.rowCount()):
            if tabla.item(row, 0).checkState() != QtCore.Qt.Checked:
                continue
            aparato = tabla.item(row, 1).text()
            cantidad = float(tabla.item(row, 2).text() or 0)
            carga_w = float(tabla.item(row, 3).text() or 0)
            horas_dia = float(tabla.item(row, 4).text() or 0)
            horas_noche = float(tabla.item(row, 5).text() or 0)
            cargas.append(
                {
                    "aparato": aparato,
                    "cantidad": cantidad,
                    "carga": carga_w,
                    "horas_dia": horas_dia,
                    "horas_noche": horas_noche,
                }
            )

        curva = curva_irradiacion_cusco()
        pot_panel, cap_bat = calcular_necesidades(cargas, curva)
        demanda_max = potencia_maxima_demanda(cargas)
        resultados = calcular_kit(datos, pot_panel, cap_bat, demanda_max)
        daily_kwh = energia_diaria_kwh(cargas, curva)
        cap_gel = cap_bat / 0.5
        cap_li = cap_bat / 0.9
        texto = (
            f"Consumo diario: {daily_kwh:.2f} kWh\n"
            f"Potencia de panel requerida: {pot_panel:.2f} W\n"
            "Capacidad de bateria requerida:\n"
            f"  Si es de Gel/Agm : {cap_gel:.2f} Ah\n"
            f"  Si es de litio: {cap_li:.2f} Ah"
        )
        salida.setPlainText(texto)

        # Graficos para la categoria Barato por defecto
        pres = resultados[CATEGORIES[0]]
        costo, _, _, _ = calcular_amortizacion(pres, daily_kwh)
        graficar_costo_acumulado(costo, daily_kwh, "resultado")
        graficar_costo_anual(costo, daily_kwh, "resultado")
        graficar_ahorro_largo_plazo(costo, daily_kwh, "resultado")

        for b in (btn_costo, btn_anual, btn_ahorro, btn_sistemas):
            b.setEnabled(True)

        mostrar_sistemas()

    def mostrar_imagen(ruta: str) -> None:
        dlg = QtWidgets.QDialog(ventana)
        dlg.resize(600, 400)
        lbl = QtWidgets.QLabel()
        pix = QtGui.QPixmap(ruta)
        lbl.setPixmap(pix)
        lay = QtWidgets.QVBoxLayout(dlg)
        lay.addWidget(lbl)
        dlg.exec_()

    def mostrar_sistemas() -> None:
        dlg = QtWidgets.QDialog(ventana)
        dlg.setWindowTitle("Sistemas recomendados")
        dlg.resize(900, 600)
        lay = QtWidgets.QVBoxLayout(dlg)

        colores = {"Barato": "#dff0d8", "Intermedio": "#fff3cd", "Premium": "#f8d7da"}
        for cat in CATEGORIES:
            pres = resultados.get(cat, {})
            if not pres:
                continue
            costo, costo_kwh, payback, ahorro = calcular_amortizacion(pres, daily_kwh)
            color = colores.get(cat, "#ffffff")
            html = f"<h3 style='background:{color};padding:4px;'>{cat}</h3>"
            html += "<table border='1' cellspacing='0' cellpadding='4' width='100%' style='margin-bottom:10px;'>"
            html += "<tr><th>Componente</th><th>Detalle</th><th>Precio</th><th>Con igv</th></tr>"
            for comp, (desc, precio) in pres.items():
                con_igv = precio * 1.18
                html += f"<tr><td>{comp}</td><td>{desc}</td><td>S/.{precio:.2f}</td><td>S/.{con_igv:.2f}</td></tr>"
            con_igv_total = costo * 1.18
            html += f"<tr style='font-weight:bold;'><td colspan='2'>Total</td><td>S/.{costo:.2f}</td><td>S/.{con_igv_total:.2f}</td></tr>"
            html += f"<tr><td colspan='2'>Costo kWh</td><td colspan='2'>S/.{costo_kwh:.2f}</td></tr>"
            html += f"<tr><td colspan='2'>Payback</td><td colspan='2'>{payback:.2f} años</td></tr>"
            html += f"<tr><td colspan='2'>Ahorro {VIDA_UTIL_ANIOS} años</td><td colspan='2'>S/.{ahorro:.2f}</td></tr>"
            html += "</table>"

            txt = QtWidgets.QTextBrowser()
            txt.setHtml(html)
            lay.addWidget(txt)
            botones = QtWidgets.QHBoxLayout()
            btn_v = QtWidgets.QPushButton(f"Vender {cat} sin IGV")
            btn_vi = QtWidgets.QPushButton(f"Vender {cat} con IGV")
            btn_v.clicked.connect(lambda _, c=cat: vender_sistema(c, False))
            btn_vi.clicked.connect(lambda _, c=cat: vender_sistema(c, True))
            botones.addWidget(btn_v)
            botones.addWidget(btn_vi)
            lay.addLayout(botones)
        dlg.exec_()

    btn_toggle.clicked.connect(toggle_checks)
    btn_ejecutar.clicked.connect(ejecutar)
    btn_costo.clicked.connect(lambda: mostrar_imagen("costo_resultado.png"))
    btn_anual.clicked.connect(lambda: mostrar_imagen("costo_anual_resultado.png"))
    btn_ahorro.clicked.connect(lambda: mostrar_imagen("ahorro_resultado.png"))
    btn_sistemas.clicked.connect(mostrar_sistemas)
    btn_inventario.clicked.connect(mostrar_inventario)
    btn_movimientos.clicked.connect(mostrar_movimientos)

    ventana.show()
    app.exec_()


if __name__ == "__main__":
    main()
