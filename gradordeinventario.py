from __future__ import annotations

import os
from typing import Dict

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover - dependency may not be installed
    Workbook = None  # type: ignore
    load_workbook = None  # type: ignore

from Precios import FILE, crear_excel_de_ejemplo

INVENTARIO_OUT = "inventario.xlsx"
TIPOS = ("Paneles", "Inversores", "Controladores")
DEFAULT_STOCK = 10


def generar_inventario(excel_file: str) -> Dict[str, Dict[str, int]]:
    """Lee el Excel de equipos y genera un inventario base."""

    if load_workbook is None:
        raise ImportError("openpyxl no esta instalado")

    if not os.path.exists(excel_file):
        crear_excel_de_ejemplo(excel_file)

    wb = load_workbook(excel_file)
    inventario: Dict[str, Dict[str, int]] = {t: {} for t in (*TIPOS, "Otros")}

    for nombre in wb.sheetnames:
        ws = wb[nombre]
        categoria = nombre if nombre in TIPOS else "Otros"
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            if len(row) < 3:
                continue
            _, marca, detalle, *_ = row
            if marca is None or detalle is None:
                continue
            producto = f"{marca} {detalle}"
            inventario[categoria][producto] = DEFAULT_STOCK
    return inventario


def guardar_inventario(filename: str, inventario: Dict[str, Dict[str, int]]) -> None:
    """Guarda el inventario en un archivo Excel con una hoja por categoria."""

    if Workbook is None:
        raise ImportError("openpyxl no esta instalado")

    wb = Workbook()
    wb.remove(wb.active)
    for categoria, items in inventario.items():
        ws = wb.create_sheet(title=categoria)
        ws.append(["Producto", "Cantidad"])
        for producto, cantidad in items.items():
            ws.append([producto, cantidad])
    wb.save(filename)


def ingresar_stock(inventario: Dict[str, Dict[str, int]], categoria: str, producto: str, cantidad: int) -> None:
    """Aumenta el stock del producto indicado."""

    if cantidad < 0:
        raise ValueError("No se puede ingresar una cantidad negativa")

    inventario.setdefault(categoria, {})
    inventario[categoria][producto] = inventario[categoria].get(producto, 0) + cantidad


def egresar_stock(inventario: Dict[str, Dict[str, int]], categoria: str, producto: str, cantidad: int) -> None:
    """Reduce el stock del producto indicado."""

    if cantidad < 0:
        raise ValueError("No se puede egresar una cantidad negativa")

    inventario.setdefault(categoria, {})
    actual = inventario[categoria].get(producto, 0)
    if actual < cantidad:
        raise ValueError("Stock insuficiente")
    inventario[categoria][producto] = actual - cantidad


def main() -> None:
    inventario = generar_inventario(FILE)
    guardar_inventario(INVENTARIO_OUT, inventario)
    print(f"Inventario guardado en '{INVENTARIO_OUT}'")


if __name__ == "__main__":
    main()
